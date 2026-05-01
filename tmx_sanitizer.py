from __future__ import annotations

import os
import re
import tempfile
import unicodedata
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime
from io import BytesIO
from typing import Dict, List, Optional, Tuple
import xml.etree.ElementTree as ET

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# App metadata
# ============================================================

APP_TITLE = "TMX Sanitizer Pro"
MAKER_LINE = "Made by LangOps Solutions"
APP_VERSION = "0.1.0"

PAGE_SIZE_DEFAULT = 200

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
ISSUE_FILL = PatternFill(fill_type="solid", fgColor="FCE5CD")
CRITICAL_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
OK_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")


# ============================================================
# Data model
# ============================================================

@dataclass
class SegmentRecord:
    record_id: int
    file_name: str
    tu_index: int
    tuid: str
    source_lang: str
    target_lang: str
    source_text: str
    target_text: str
    creation_date: str = ""
    change_date: str = ""
    creation_id: str = ""
    change_id: str = ""

    source_length: int = 0
    target_length: int = 0

    issue_count: int = 0
    severity: str = "OK"
    issue_categories: str = ""
    issue_details: str = ""

    repair_actions: str = ""


@dataclass
class SanitizerSettings:
    preferred_source_lang: str = ""
    preferred_target_lang: str = ""
    normalize_unicode: bool = True
    trim_spaces: bool = True
    remove_zero_width: bool = True
    replace_nbsp: bool = True
    collapse_spaces: bool = True
    normalize_language_codes: bool = True
    flag_tag_issues: bool = True
    flag_source_equals_target: bool = True
    flag_german_micro_qa: bool = True
    flag_brand_protection: bool = True
    flag_placeholder_issues: bool = True
    flag_number_issues: bool = True
    flag_punctuation_issues: bool = True
    flag_length_ratio: bool = True


# ============================================================
# Language normalization
# ============================================================

LANGUAGE_CODE_MAP = {
    # German
    "de": "de-DE",
    "de-de": "de-DE",
    "de_de": "de-DE",
    "deu": "de-DE",
    "ger": "de-DE",
    "german": "de-DE",
    "de-at": "de-AT",
    "de_at": "de-AT",
    "de-ch": "de-CH",
    "de_ch": "de-CH",

    # English
    "en": "en-US",
    "en-us": "en-US",
    "en_us": "en-US",
    "en-us.": "en-US",
    "eng": "en-US",
    "english": "en-US",
    "en-gb": "en-GB",
    "en_gb": "en-GB",
    "en-uk": "en-GB",
    "en_uk": "en-GB",

    # French
    "fr": "fr-FR",
    "fr-fr": "fr-FR",
    "fr_fr": "fr-FR",
    "fre": "fr-FR",
    "fra": "fr-FR",
    "french": "fr-FR",
    "fr-ca": "fr-CA",
    "fr_ca": "fr-CA",

    # Spanish
    "es": "es-ES",
    "es-es": "es-ES",
    "es_es": "es-ES",
    "spa": "es-ES",
    "spanish": "es-ES",
    "es-mx": "es-MX",
    "es_mx": "es-MX",
    "es-ww": "es-WW",
    "es_ww": "es-WW",

    # Portuguese
    "pt": "pt-PT",
    "pt-pt": "pt-PT",
    "pt_pt": "pt-PT",
    "pt-br": "pt-BR",
    "pt_br": "pt-BR",
    "por": "pt-PT",

    # Others
    "it": "it-IT",
    "it-it": "it-IT",
    "it_it": "it-IT",
    "ita": "it-IT",
    "nl": "nl-NL",
    "nl-nl": "nl-NL",
    "nl_nl": "nl-NL",
    "dut": "nl-NL",
    "nld": "nl-NL",
    "ja": "ja-JP",
    "ja-jp": "ja-JP",
    "ja_jp": "ja-JP",
    "jpn": "ja-JP",
    "ko": "ko-KR",
    "ko-kr": "ko-KR",
    "ko_kr": "ko-KR",
    "kor": "ko-KR",
    "zh": "zh-CN",
    "zh-cn": "zh-CN",
    "zh_cn": "zh-CN",
    "zh-hans": "zh-CN",
    "zh-tw": "zh-TW",
    "zh_tw": "zh-TW",
    "zh-hant": "zh-TW",
    "ar": "ar-SA",
    "ar-sa": "ar-SA",
    "ar_sa": "ar-SA",
    "pl": "pl-PL",
    "pl-pl": "pl-PL",
    "pl_pl": "pl-PL",
    "cs": "cs-CZ",
    "cs-cz": "cs-CZ",
    "cs_cz": "cs-CZ",
    "sv": "sv-SE",
    "sv-se": "sv-SE",
    "sv_se": "sv-SE",
}


def normalize_language_code(code: str) -> str:
    raw = (code or "").strip()
    if not raw:
        return raw
    key = raw.lower().replace(" ", "").replace("_", "-")
    key_alt = raw.lower().replace(" ", "")
    return LANGUAGE_CODE_MAP.get(key) or LANGUAGE_CODE_MAP.get(key_alt) or raw.replace("_", "-")


def lang_is_german(code: str) -> bool:
    return (code or "").lower().startswith("de")


def lang_matches(actual: str, preferred: str) -> bool:
    if not preferred:
        return True
    actual_n = (actual or "").strip().lower().replace("_", "-")
    preferred_n = (preferred or "").strip().lower().replace("_", "-")
    return actual_n == preferred_n or actual_n.startswith(preferred_n + "-") or preferred_n.startswith(actual_n + "-")


# ============================================================
# TMX parsing
# ============================================================

class TMXParser:
    XML_LANG = "{http://www.w3.org/XML/1998/namespace}lang"

    @staticmethod
    def strip_namespace(tag: str) -> str:
        return tag.split("}", 1)[1] if "}" in tag else tag

    @staticmethod
    def xml_lang(elem: ET.Element) -> str:
        return elem.attrib.get(TMXParser.XML_LANG) or elem.attrib.get("lang") or ""

    @staticmethod
    def clean_visible_text(text: str) -> str:
        text = text or ""
        text = re.sub(r"\s+", " ", text)
        return text.strip()

    @staticmethod
    def parse_file(path: str, file_name: str, start_id: int, preferred_source: str = "", preferred_target: str = "") -> List[SegmentRecord]:
        records: List[SegmentRecord] = []
        next_id = start_id
        tu_index = 0

        try:
            context = ET.iterparse(path, events=("end",))
        except ET.ParseError as exc:
            raise ValueError(f"Could not parse TMX file '{file_name}': {exc}") from exc

        for _event, elem in context:
            if TMXParser.strip_namespace(elem.tag) != "tu":
                continue

            tu_index += 1
            tuid = elem.attrib.get("tuid", "")
            creation_date = elem.attrib.get("creationdate", "")
            change_date = elem.attrib.get("changedate", "")
            creation_id = elem.attrib.get("creationid", "")
            change_id = elem.attrib.get("changeid", "")

            tuvs: List[Tuple[str, str]] = []

            for child in list(elem):
                if TMXParser.strip_namespace(child.tag) != "tuv":
                    continue

                lang = TMXParser.xml_lang(child).strip()
                seg_text = ""

                for seg in list(child):
                    if TMXParser.strip_namespace(seg.tag) == "seg":
                        seg_text = "".join(seg.itertext())
                        break

                tuvs.append((lang, TMXParser.clean_visible_text(seg_text)))

            if not tuvs:
                elem.clear()
                continue

            source_lang = ""
            source_text = ""
            target_lang = ""
            target_text = ""

            if len(tuvs) >= 2:
                if preferred_source and preferred_target:
                    src = next((item for item in tuvs if lang_matches(item[0], preferred_source)), None)
                    tgt = next((item for item in tuvs if lang_matches(item[0], preferred_target)), None)

                    if src and tgt and src != tgt:
                        source_lang, source_text = src
                        target_lang, target_text = tgt
                    else:
                        source_lang, source_text = tuvs[0]
                        target_lang, target_text = tuvs[1]
                else:
                    source_lang, source_text = tuvs[0]
                    target_lang, target_text = tuvs[1]
            else:
                source_lang, source_text = tuvs[0]

            records.append(
                SegmentRecord(
                    record_id=next_id,
                    file_name=file_name,
                    tu_index=tu_index,
                    tuid=tuid,
                    source_lang=source_lang,
                    target_lang=target_lang,
                    source_text=source_text,
                    target_text=target_text,
                    creation_date=creation_date,
                    change_date=change_date,
                    creation_id=creation_id,
                    change_id=change_id,
                    source_length=len(source_text or ""),
                    target_length=len(target_text or ""),
                )
            )
            next_id += 1
            elem.clear()

        return records


def parse_uploaded_tmx_files(uploaded_files, preferred_source: str, preferred_target: str) -> List[SegmentRecord]:
    all_records: List[SegmentRecord] = []
    next_id = 1

    for uploaded in uploaded_files:
        suffix = os.path.splitext(uploaded.name)[1] or ".tmx"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(uploaded.getvalue())
            tmp_path = tmp.name

        try:
            records = TMXParser.parse_file(
                path=tmp_path,
                file_name=uploaded.name,
                start_id=next_id,
                preferred_source=preferred_source,
                preferred_target=preferred_target,
            )
            all_records.extend(records)
            if records:
                next_id = max(r.record_id for r in all_records) + 1
        finally:
            try:
                os.remove(tmp_path)
            except OSError:
                pass

    return all_records


# ============================================================
# Rule loaders
# ============================================================

class BrandProtectionRules:
    def __init__(self) -> None:
        self.rules: List[Dict[str, str]] = []

    def load_xlsx(self, uploaded_file) -> int:
        df = pd.read_excel(uploaded_file, header=None)
        if df.shape[1] < 2:
            raise ValueError("Do Not Translate XLSX must use Column A = source term and Column B = required target representation.")

        rules: List[Dict[str, str]] = []

        for _, row in df.iterrows():
            source = "" if pd.isna(row.iloc[0]) else str(row.iloc[0]).strip()
            required = "" if pd.isna(row.iloc[1]) else str(row.iloc[1]).strip()
            note = "" if df.shape[1] < 3 or pd.isna(row.iloc[2]) else str(row.iloc[2]).strip()

            if source and required:
                rules.append({"source": source, "required": required, "note": note})

        self.rules = rules
        return len(self.rules)


# ============================================================
# Repairs
# ============================================================

class RepairEngine:
    ZERO_WIDTH_PATTERN = re.compile(r"[\u200b\u200c\u200d\ufeff]")

    @staticmethod
    def repair_text(text: str, settings: SanitizerSettings) -> Tuple[str, List[str]]:
        new_text = text or ""
        actions: List[str] = []

        before = new_text
        if settings.normalize_unicode:
            new_text = unicodedata.normalize("NFC", new_text)
            if new_text != before:
                actions.append("Unicode normalized")

        before = new_text
        if settings.replace_nbsp:
            new_text = new_text.replace("\xa0", " ")
            if new_text != before:
                actions.append("NBSP replaced")

        before = new_text
        if settings.remove_zero_width:
            new_text = RepairEngine.ZERO_WIDTH_PATTERN.sub("", new_text)
            if new_text != before:
                actions.append("Zero-width character removed")

        before = new_text
        if settings.collapse_spaces:
            new_text = re.sub(r"[ \t]{2,}", " ", new_text)
            if new_text != before:
                actions.append("Multiple spaces collapsed")

        before = new_text
        if settings.trim_spaces:
            new_text = new_text.strip()
            if new_text != before:
                actions.append("Trimmed")

        return new_text, actions

    @staticmethod
    def repair_record(record: SegmentRecord, settings: SanitizerSettings) -> None:
        actions: List[str] = []

        source_text, src_actions = RepairEngine.repair_text(record.source_text, settings)
        target_text, tgt_actions = RepairEngine.repair_text(record.target_text, settings)

        record.source_text = source_text
        record.target_text = target_text

        actions.extend([f"Source: {a}" for a in src_actions])
        actions.extend([f"Target: {a}" for a in tgt_actions])

        if settings.normalize_language_codes:
            old_source = record.source_lang
            old_target = record.target_lang
            record.source_lang = normalize_language_code(record.source_lang)
            record.target_lang = normalize_language_code(record.target_lang)

            if record.source_lang != old_source:
                actions.append(f"Source lang {old_source} -> {record.source_lang}")
            if record.target_lang != old_target:
                actions.append(f"Target lang {old_target} -> {record.target_lang}")

        record.source_length = len(record.source_text or "")
        record.target_length = len(record.target_text or "")
        record.repair_actions = "; ".join(actions)

    @staticmethod
    def repair_all(records: List[SegmentRecord], settings: SanitizerSettings) -> int:
        changed = 0
        for record in records:
            before = (
                record.source_text,
                record.target_text,
                record.source_lang,
                record.target_lang,
            )
            RepairEngine.repair_record(record, settings)
            after = (
                record.source_text,
                record.target_text,
                record.source_lang,
                record.target_lang,
            )
            if before != after:
                changed += 1
        return changed


# ============================================================
# QA / flagging
# ============================================================

class QAEngine:
    PLACEHOLDER_PATTERNS = [
        r"\{\d+\}",
        r"\{[A-Za-z0-9_]+\}",
        r"%s",
        r"%d",
        r"\$\{[A-Za-z0-9_]+\}",
        r"<[^>]+>",
    ]

    TAG_PATTERN = re.compile(r"</?([A-Za-z][A-Za-z0-9:_-]*)(?:\s[^>]*)?>")
    SELF_CLOSING_PATTERN = re.compile(r"<([A-Za-z][A-Za-z0-9:_-]*)(?:\s[^>]*)?/>")

    @staticmethod
    def extract_placeholders(text: str) -> List[str]:
        found: List[str] = []
        for pattern in QAEngine.PLACEHOLDER_PATTERNS:
            found.extend(re.findall(pattern, text or ""))
        return sorted(found)

    @staticmethod
    def extract_numbers(text: str) -> List[str]:
        return re.findall(r"\d+(?:[\.,]\d+)?", text or "")

    @staticmethod
    def end_punctuation(text: str) -> str:
        text = (text or "").strip()
        return text[-1] if text and text[-1] in ".,:;!?" else ""

    @staticmethod
    def has_tag_balance_issue(text: str) -> bool:
        text = text or ""
        if "<" not in text and ">" not in text:
            return False

        if text.count("<") != text.count(">"):
            return True

        stack: List[str] = []
        for match in QAEngine.TAG_PATTERN.finditer(text):
            full = match.group(0)
            tag = match.group(1).lower()

            if full.endswith("/>"):
                continue

            if full.startswith("</"):
                if not stack or stack[-1] != tag:
                    return True
                stack.pop()
            else:
                # Ignore common standalone HTML-like tags.
                if tag in {"br", "hr", "img", "input", "meta", "link"}:
                    continue
                stack.append(tag)

        return bool(stack)

    @staticmethod
    def brand_violations(source: str, target: str, rules: BrandProtectionRules) -> List[str]:
        violations: List[str] = []
        source_l = source or ""
        target_l = target or ""

        for item in rules.rules:
            src = item["source"]
            required = item["required"]

            pattern = r"(?<!\w)" + re.escape(src) + r"(?!\w)"
            if re.search(pattern, source_l, flags=re.IGNORECASE):
                required_pattern = r"(?<!\w)" + re.escape(required) + r"(?!\w)"
                if not re.search(required_pattern, target_l, flags=re.IGNORECASE):
                    violations.append(f"Protected term '{src}' should appear as '{required}'")

        return violations

    @staticmethod
    def german_micro_issues(record: SegmentRecord) -> List[str]:
        issues: List[str] = []
        if not lang_is_german(record.target_lang):
            return issues

        target = record.target_text or ""

        if re.search(r"\s+[.,:;!?]", target):
            issues.append("German QA: space before punctuation")

        if re.search(r"\b(\w+)\s+\1\b", target, flags=re.IGNORECASE):
            issues.append("German QA: repeated word")

        if re.search(r"\d+\s?%", target):
            issues.append("German QA: check percent spacing style")

        if '"' in target:
            issues.append("German QA: straight quotes used; check German quotation style")

        if re.search(r"\b[A-Z][a-z]+ing\b", target):
            issues.append("German QA: possible untranslated English -ing form")

        if re.search(r"\b(the|and|with|for|from|your|our|you)\b", target, flags=re.IGNORECASE):
            issues.append("German QA: possible English word in German target")

        return issues

    @staticmethod
    def apply(records: List[SegmentRecord], settings: SanitizerSettings, brand_rules: BrandProtectionRules) -> None:
        for record in records:
            issues: List[str] = []
            categories: List[str] = []

            source = record.source_text or ""
            target = record.target_text or ""

            if not target.strip():
                issues.append("Missing target")
                categories.append("Missing Target")

            if settings.flag_source_equals_target and source.strip() and target.strip() and source.strip() == target.strip():
                issues.append("Target equals source; possible untranslated segment")
                categories.append("Source=Target")

            if settings.flag_placeholder_issues:
                if QAEngine.extract_placeholders(source) != QAEngine.extract_placeholders(target):
                    issues.append("Placeholder/tag mismatch")
                    categories.append("Placeholders")

            if settings.flag_number_issues:
                if QAEngine.extract_numbers(source) != QAEngine.extract_numbers(target):
                    issues.append("Number mismatch")
                    categories.append("Numbers")

            if settings.flag_punctuation_issues:
                src_end = QAEngine.end_punctuation(source)
                tgt_end = QAEngine.end_punctuation(target)
                if src_end != tgt_end and (src_end or tgt_end):
                    issues.append("Ending punctuation mismatch")
                    categories.append("Punctuation")

            if settings.flag_length_ratio and source:
                ratio = len(target) / max(1, len(source))
                if ratio < 0.35 or ratio > 2.8:
                    issues.append("Suspicious source/target length ratio")
                    categories.append("Length")

            if settings.flag_tag_issues:
                if QAEngine.has_tag_balance_issue(source):
                    issues.append("Source contains unbalanced or malformed HTML/XML-like tags")
                    categories.append("Tags")
                if QAEngine.has_tag_balance_issue(target):
                    issues.append("Target contains unbalanced or malformed HTML/XML-like tags")
                    categories.append("Tags")

            if settings.flag_brand_protection and brand_rules.rules:
                brand_hits = QAEngine.brand_violations(source, target, brand_rules)
                if brand_hits:
                    issues.extend(brand_hits)
                    categories.append("Brand Protection")

            if settings.flag_german_micro_qa:
                german_hits = QAEngine.german_micro_issues(record)
                if german_hits:
                    issues.extend(german_hits)
                    categories.append("German Micro QA")

            unique_categories = sorted(set(categories))
            record.issue_count = len(issues)
            record.issue_categories = "; ".join(unique_categories)
            record.issue_details = "; ".join(issues)
            record.severity = "Issues" if issues else "OK"


# ============================================================
# Duplicate and merge logic
# ============================================================

class DuplicateEngine:
    @staticmethod
    def normalize_source(text: str) -> str:
        text = unicodedata.normalize("NFC", text or "")
        text = re.sub(r"\s+", " ", text.strip().lower())
        text = re.sub(r"[\W_]+", "", text, flags=re.UNICODE)
        return text

    @staticmethod
    def duplicate_summary(records: List[SegmentRecord]) -> Dict[str, int]:
        exact_pairs = Counter((r.source_lang, r.target_lang, r.source_text, r.target_text) for r in records)
        source_targets: Dict[Tuple[str, str], set] = defaultdict(set)
        normalized = Counter((r.source_lang, DuplicateEngine.normalize_source(r.source_text)) for r in records)

        for r in records:
            source_targets[(r.source_lang, r.source_text)].add(r.target_text)

        return {
            "exact_source_target_duplicates": sum(1 for r in records if exact_pairs[(r.source_lang, r.target_lang, r.source_text, r.target_text)] > 1),
            "same_source_different_target": sum(1 for r in records if len(source_targets[(r.source_lang, r.source_text)]) > 1),
            "normalized_source_duplicates": sum(1 for r in records if normalized[(r.source_lang, DuplicateEngine.normalize_source(r.source_text))] > 1),
        }


class MergeEngine:
    @staticmethod
    def parse_date(value: str) -> datetime:
        value = (value or "").strip()
        if not value:
            return datetime.min

        for fmt in ["%Y%m%dT%H%M%SZ", "%Y%m%dT%H%M%S", "%Y%m%d", "%Y-%m-%d"]:
            try:
                return datetime.strptime(value, fmt)
            except Exception:
                pass

        return datetime.min

    @staticmethod
    def choose_best(records: List[SegmentRecord], strategy: str) -> SegmentRecord:
        if strategy == "Keep newest translation":
            return max(records, key=lambda r: (MergeEngine.parse_date(r.change_date), MergeEngine.parse_date(r.creation_date), len(r.target_text or "")))

        if strategy == "Keep longest translation":
            return max(records, key=lambda r: (len(r.target_text or ""), len(r.source_text or "")))

        if strategy == "Keep lowest issue count":
            return min(records, key=lambda r: (r.issue_count, -len(r.target_text or "")))

        return records[0]

    @staticmethod
    def merge(records: List[SegmentRecord], strategy: str) -> Tuple[List[SegmentRecord], int]:
        grouped: Dict[Tuple[str, str, str], List[SegmentRecord]] = defaultdict(list)
        for r in records:
            grouped[(r.source_lang, r.target_lang, r.source_text)].append(r)

        merged: List[SegmentRecord] = []
        removed = 0

        for group in grouped.values():
            if len(group) == 1:
                merged.append(group[0])
            else:
                merged.append(MergeEngine.choose_best(group, strategy))
                removed += len(group) - 1

        merged.sort(key=lambda r: (r.file_name, r.tu_index, r.record_id))
        return merged, removed


# ============================================================
# Exporters
# ============================================================

class TMXExporter:
    @staticmethod
    def to_bytes(records: List[SegmentRecord]) -> bytes:
        root = ET.Element("tmx", version="1.4")
        ET.SubElement(
            root,
            "header",
            {
                "creationtool": MAKER_LINE,
                "creationtoolversion": APP_VERSION,
                "segtype": "sentence",
                "adminlang": "en-US",
                "srclang": records[0].source_lang if records else "en-US",
                "datatype": "PlainText",
                "creationdate": datetime.utcnow().strftime("%Y%m%dT%H%M%SZ"),
            },
        )

        body = ET.SubElement(root, "body")

        for idx, record in enumerate(records, start=1):
            tu_attrib = {"tuid": record.tuid or str(idx)}
            if record.creation_date:
                tu_attrib["creationdate"] = record.creation_date
            if record.change_date:
                tu_attrib["changedate"] = record.change_date
            if record.creation_id:
                tu_attrib["creationid"] = record.creation_id
            if record.change_id:
                tu_attrib["changeid"] = record.change_id

            tu = ET.SubElement(body, "tu", tu_attrib)

            tuv_source = ET.SubElement(tu, "tuv", {"{http://www.w3.org/XML/1998/namespace}lang": record.source_lang or "en-US"})
            ET.SubElement(tuv_source, "seg").text = record.source_text

            tuv_target = ET.SubElement(tu, "tuv", {"{http://www.w3.org/XML/1998/namespace}lang": record.target_lang or "de-DE"})
            ET.SubElement(tuv_target, "seg").text = record.target_text

        output = BytesIO()
        ET.ElementTree(root).write(output, encoding="utf-8", xml_declaration=True)
        return output.getvalue()


class XLSXReportExporter:
    COLUMNS = [
        "Record ID",
        "File",
        "TU Index",
        "TU ID",
        "Source Lang",
        "Target Lang",
        "Source",
        "Target",
        "Severity",
        "Issue Count",
        "Issue Categories",
        "Issue Details",
        "Repair Actions",
        "Source Length",
        "Target Length",
        "Creation Date",
        "Change Date",
        "Creation ID",
        "Change ID",
    ]

    @staticmethod
    def append_record(ws, r: SegmentRecord) -> None:
        ws.append([
            r.record_id,
            r.file_name,
            r.tu_index,
            r.tuid,
            r.source_lang,
            r.target_lang,
            r.source_text,
            r.target_text,
            r.severity,
            r.issue_count,
            r.issue_categories,
            r.issue_details,
            r.repair_actions,
            r.source_length,
            r.target_length,
            r.creation_date,
            r.change_date,
            r.creation_id,
            r.change_id,
        ])

    @staticmethod
    def style_sheet(ws) -> None:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.freeze_panes = "A2"

        for row in range(2, ws.max_row + 1):
            severity = ws.cell(row=row, column=9).value
            issue_count = ws.cell(row=row, column=10).value or 0

            if severity == "Issues" and issue_count >= 3:
                fill = CRITICAL_FILL
            elif severity == "Issues":
                fill = ISSUE_FILL
            else:
                fill = OK_FILL

            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill
                ws.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)

        for col in ws.columns:
            max_len = 0
            letter = get_column_letter(col[0].column)
            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            ws.column_dimensions[letter].width = min(max_len + 2, 80)

    @staticmethod
    def to_bytes(records: List[SegmentRecord], stats: Dict[str, object]) -> bytes:
        wb = Workbook()

        ws = wb.active
        ws.title = "Segments"
        ws.append(XLSXReportExporter.COLUMNS)
        for record in records:
            XLSXReportExporter.append_record(ws, record)
        XLSXReportExporter.style_sheet(ws)

        issue_ws = wb.create_sheet("Issues Only")
        issue_ws.append(XLSXReportExporter.COLUMNS)
        for record in records:
            if record.issue_count:
                XLSXReportExporter.append_record(issue_ws, record)
        XLSXReportExporter.style_sheet(issue_ws)

        repair_ws = wb.create_sheet("Repair Actions")
        repair_ws.append(XLSXReportExporter.COLUMNS)
        for record in records:
            if record.repair_actions:
                XLSXReportExporter.append_record(repair_ws, record)
        XLSXReportExporter.style_sheet(repair_ws)

        stats_ws = wb.create_sheet("Statistics")
        stats_ws.append(["Metric", "Value"])
        for key, value in stats.items():
            if isinstance(value, (dict, Counter)):
                stats_ws.append([key, ""])
                for sub_key, sub_value in value.items():
                    stats_ws.append([f"  {sub_key}", sub_value])
            else:
                stats_ws.append([key, value])
        XLSXReportExporter.style_sheet(stats_ws)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


# ============================================================
# State / utility
# ============================================================

def init_state() -> None:
    if "records" not in st.session_state:
        st.session_state.records = []
    if "stats" not in st.session_state:
        st.session_state.stats = {}
    if "brand_rules" not in st.session_state:
        st.session_state.brand_rules = BrandProtectionRules()
    if "logs" not in st.session_state:
        st.session_state.logs = []


def log(message: str) -> None:
    timestamp = datetime.now().strftime("%H:%M:%S")
    st.session_state.logs.append(f"[{timestamp}] {message}")


def records_to_dataframe(records: List[SegmentRecord]) -> pd.DataFrame:
    rows = []
    for r in records:
        rows.append({
            "Record ID": r.record_id,
            "File": r.file_name,
            "TU Index": r.tu_index,
            "Source Lang": r.source_lang,
            "Target Lang": r.target_lang,
            "Source": r.source_text,
            "Target": r.target_text,
            "Severity": r.severity,
            "Issue Count": r.issue_count,
            "Issue Categories": r.issue_categories,
            "Issue Details": r.issue_details,
            "Repair Actions": r.repair_actions,
        })
    return pd.DataFrame(rows)


def build_stats(records: List[SegmentRecord]) -> Dict[str, object]:
    dup = DuplicateEngine.duplicate_summary(records)
    return {
        "Tool": f"{APP_TITLE} {APP_VERSION}",
        "Maker": MAKER_LINE,
        "Total Segments": len(records),
        "Segments with Issues": sum(1 for r in records if r.issue_count),
        "Clean Segments": sum(1 for r in records if not r.issue_count),
        "Average Source Length": round(sum(r.source_length for r in records) / max(1, len(records)), 2),
        "Average Target Length": round(sum(r.target_length for r in records) / max(1, len(records)), 2),
        "Language Pairs": Counter(f"{r.source_lang} > {r.target_lang}" for r in records),
        "Issue Categories": Counter(cat.strip() for r in records for cat in r.issue_categories.split(";") if cat.strip()),
        "Duplicate Summary": dup,
        "Per File Counts": Counter(r.file_name for r in records),
    }


def get_settings_from_sidebar() -> SanitizerSettings:
    st.sidebar.header("Sanitizer Settings")

    preferred_source = st.sidebar.text_input("Preferred source language", value="", placeholder="e.g. en, en-US")
    preferred_target = st.sidebar.text_input("Preferred target language", value="", placeholder="e.g. de, de-DE")

    st.sidebar.subheader("Safe Auto-Repairs")
    normalize_unicode = st.sidebar.checkbox("Normalize Unicode NFC", value=True)
    trim_spaces = st.sidebar.checkbox("Trim leading/trailing spaces", value=True)
    remove_zero_width = st.sidebar.checkbox("Remove zero-width characters", value=True)
    replace_nbsp = st.sidebar.checkbox("Replace non-breaking spaces", value=True)
    collapse_spaces = st.sidebar.checkbox("Collapse repeated spaces", value=True)
    normalize_langs = st.sidebar.checkbox("Normalize language codes", value=True)

    st.sidebar.subheader("Flag-Only QA Checks")
    flag_tags = st.sidebar.checkbox("Flag malformed/unbalanced tags", value=True)
    flag_source_equals_target = st.sidebar.checkbox("Flag target equals source", value=True)
    flag_german_micro = st.sidebar.checkbox("Flag German micro-QA issues", value=True)
    flag_brand = st.sidebar.checkbox("Flag brand / do-not-translate issues", value=True)
    flag_placeholders = st.sidebar.checkbox("Flag placeholder mismatch", value=True)
    flag_numbers = st.sidebar.checkbox("Flag number mismatch", value=True)
    flag_punctuation = st.sidebar.checkbox("Flag punctuation mismatch", value=True)
    flag_length = st.sidebar.checkbox("Flag suspicious length ratio", value=True)

    return SanitizerSettings(
        preferred_source_lang=preferred_source.strip(),
        preferred_target_lang=preferred_target.strip(),
        normalize_unicode=normalize_unicode,
        trim_spaces=trim_spaces,
        remove_zero_width=remove_zero_width,
        replace_nbsp=replace_nbsp,
        collapse_spaces=collapse_spaces,
        normalize_language_codes=normalize_langs,
        flag_tag_issues=flag_tags,
        flag_source_equals_target=flag_source_equals_target,
        flag_german_micro_qa=flag_german_micro,
        flag_brand_protection=flag_brand,
        flag_placeholder_issues=flag_placeholders,
        flag_number_issues=flag_numbers,
        flag_punctuation_issues=flag_punctuation,
        flag_length_ratio=flag_length,
    )


def apply_filters(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    with st.expander("Filters", expanded=True):
        col1, col2, col3 = st.columns(3)
        with col1:
            severity = st.selectbox("Severity", ["All", "Issues", "OK"])
        with col2:
            categories = sorted({cat.strip() for val in df["Issue Categories"].fillna("") for cat in str(val).split(";") if cat.strip()})
            category = st.selectbox("Issue Category", ["All"] + categories)
        with col3:
            search = st.text_input("Search source/target/file")

    filtered = df.copy()

    if severity != "All":
        filtered = filtered[filtered["Severity"] == severity]

    if category != "All":
        filtered = filtered[filtered["Issue Categories"].fillna("").str.contains(re.escape(category), case=False, regex=True)]

    if search.strip():
        needle = search.strip().lower()
        mask = (
            filtered["File"].fillna("").str.lower().str.contains(re.escape(needle), regex=True)
            | filtered["Source"].fillna("").str.lower().str.contains(re.escape(needle), regex=True)
            | filtered["Target"].fillna("").str.lower().str.contains(re.escape(needle), regex=True)
            | filtered["Issue Details"].fillna("").str.lower().str.contains(re.escape(needle), regex=True)
        )
        filtered = filtered[mask]

    return filtered


def rerun_qa(settings: SanitizerSettings) -> None:
    QAEngine.apply(st.session_state.records, settings, st.session_state.brand_rules)
    st.session_state.stats = build_stats(st.session_state.records)


# ============================================================
# Streamlit app
# ============================================================

def main() -> None:
    st.set_page_config(page_title=APP_TITLE, page_icon="🧼", layout="wide")
    init_state()

    st.title("🧼 TMX Sanitizer Pro")
    st.caption("Enterprise-style TMX cleanup, QA flagging, brand protection, and controlled export.")

    settings = get_settings_from_sidebar()

    st.sidebar.divider()
    st.sidebar.subheader("Upload Rule Files")

    brand_file = st.sidebar.file_uploader(
        "Do Not Translate / Brand Protection XLSX",
        type=["xlsx"],
        help="Column A = source protected term, Column B = required target representation, optional Column C = note.",
    )

    if brand_file is not None:
        try:
            count = st.session_state.brand_rules.load_xlsx(brand_file)
            st.sidebar.success(f"Loaded {count} protected term rules.")
        except Exception as exc:
            st.sidebar.error(str(exc))

    st.sidebar.divider()
    st.sidebar.caption(f"{MAKER_LINE} · v{APP_VERSION}")

    tab_upload, tab_dashboard, tab_segments, tab_duplicates, tab_edit, tab_export, tab_logs = st.tabs(
        ["Upload & Analyze", "Dashboard", "Segments", "Duplicates", "Edit", "Export", "Logs"]
    )

    with tab_upload:
        st.subheader("Upload TMX Files")
        uploaded_tmx = st.file_uploader(
            "Upload one or multiple TMX files",
            type=["tmx"],
            accept_multiple_files=True,
        )

        col1, col2, col3 = st.columns(3)

        with col1:
            if st.button("Analyze TMX Files", type="primary", use_container_width=True):
                if not uploaded_tmx:
                    st.warning("Please upload at least one TMX file.")
                else:
                    try:
                        with st.spinner("Parsing TMX files and running QA..."):
                            records = parse_uploaded_tmx_files(
                                uploaded_tmx,
                                preferred_source=settings.preferred_source_lang,
                                preferred_target=settings.preferred_target_lang,
                            )
                            RepairEngine.repair_all(records, settings)
                            st.session_state.records = records
                            rerun_qa(settings)
                            log(f"Analyzed {len(uploaded_tmx)} TMX file(s), {len(records)} segment(s).")
                        st.success(f"Analysis complete: {len(records)} segments loaded.")
                    except Exception as exc:
                        st.error(str(exc))
                        log(f"Analysis error: {exc}")

        with col2:
            if st.button("Run Selected Auto-Repairs", use_container_width=True):
                if not st.session_state.records:
                    st.warning("Please analyze TMX files first.")
                else:
                    changed = RepairEngine.repair_all(st.session_state.records, settings)
                    rerun_qa(settings)
                    log(f"Ran selected auto-repairs. Changed {changed} record(s).")
                    st.success(f"Auto-repairs complete. Changed {changed} record(s).")

        with col3:
            if st.button("Clear Project", use_container_width=True):
                st.session_state.records = []
                st.session_state.stats = {}
                st.session_state.logs = []
                st.session_state.brand_rules = BrandProtectionRules()
                st.success("Project cleared.")

        st.info(
            "Safe repairs are controlled by the sidebar checkboxes. Risky items such as tag imbalance, brand violations, "
            "German micro-QA issues, and target-equals-source are flagged only."
        )

    with tab_dashboard:
        records = st.session_state.records
        stats = st.session_state.stats

        if not records:
            st.info("Upload and analyze TMX files to see the dashboard.")
        else:
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Total Segments", stats.get("Total Segments", 0))
            c2.metric("Segments with Issues", stats.get("Segments with Issues", 0))
            c3.metric("Clean Segments", stats.get("Clean Segments", 0))
            c4.metric("Protected Terms", len(st.session_state.brand_rules.rules))

            st.divider()

            col1, col2 = st.columns(2)

            with col1:
                issue_counter = stats.get("Issue Categories", Counter())
                if issue_counter:
                    st.subheader("Issue Categories")
                    chart_df = pd.DataFrame(issue_counter.items(), columns=["Issue Category", "Count"])
                    st.bar_chart(chart_df, x="Issue Category", y="Count")
                else:
                    st.success("No issue categories found.")

            with col2:
                lang_pairs = stats.get("Language Pairs", Counter())
                if lang_pairs:
                    st.subheader("Language Pairs")
                    lp_df = pd.DataFrame(lang_pairs.items(), columns=["Language Pair", "Count"])
                    st.dataframe(lp_df, use_container_width=True, hide_index=True)

            st.subheader("Duplicate Summary")
            st.json(stats.get("Duplicate Summary", {}))

    with tab_segments:
        records = st.session_state.records

        if not records:
            st.info("No records loaded.")
        else:
            df = records_to_dataframe(records)
            filtered = apply_filters(df)

            st.caption(f"Showing {len(filtered)} of {len(df)} records.")
            st.dataframe(filtered, use_container_width=True, hide_index=True, height=600)

    with tab_duplicates:
        records = st.session_state.records

        if not records:
            st.info("No records loaded.")
        else:
            st.subheader("Duplicate Analysis")

            source_groups: Dict[Tuple[str, str, str], List[SegmentRecord]] = defaultdict(list)
            for r in records:
                source_groups[(r.source_lang, r.target_lang, r.source_text)].append(r)

            dup_groups = {k: v for k, v in source_groups.items() if len(v) > 1}

            st.caption(f"Found {len(dup_groups)} same-source duplicate group(s).")

            if dup_groups:
                group_labels = [
                    f"{idx + 1}: {key[0]}>{key[1]} | {key[2][:80]} ({len(items)} records)"
                    for idx, (key, items) in enumerate(dup_groups.items())
                ]
                selected_label = st.selectbox("Duplicate Group", group_labels)
                selected_index = group_labels.index(selected_label)
                selected_items = list(dup_groups.values())[selected_index]

                st.dataframe(records_to_dataframe(selected_items), use_container_width=True, hide_index=True)

                keep_id = st.selectbox("Record ID to keep", [r.record_id for r in selected_items])

                if st.button("Keep Selected / Remove Others"):
                    delete_ids = {r.record_id for r in selected_items if r.record_id != keep_id}
                    st.session_state.records = [r for r in st.session_state.records if r.record_id not in delete_ids]
                    rerun_qa(settings)
                    log(f"Duplicate resolver kept {keep_id} and removed {len(delete_ids)} record(s).")
                    st.success(f"Removed {len(delete_ids)} duplicate record(s).")
                    st.rerun()

            st.divider()
            st.subheader("Merge Policy")

            strategy = st.selectbox(
                "Merge strategy",
                ["Keep newest translation", "Keep longest translation", "Keep lowest issue count", "Keep first occurrence"],
            )

            if st.button("Apply Merge Policy to All Same-Source Groups"):
                merged, removed = MergeEngine.merge(st.session_state.records, strategy)
                st.session_state.records = merged
                rerun_qa(settings)
                log(f"Applied merge policy '{strategy}'. Removed {removed} record(s).")
                st.success(f"Merge complete. Removed {removed} record(s).")
                st.rerun()

    with tab_edit:
        records = st.session_state.records

        if not records:
            st.info("No records loaded.")
        else:
            st.subheader("Human-in-the-Loop Segment Editing")

            record_ids = [r.record_id for r in records]
            selected_id = st.selectbox("Select Record ID", record_ids)

            record = next(r for r in records if r.record_id == selected_id)

            with st.form("edit_record_form"):
                col1, col2, col3 = st.columns(3)
                with col1:
                    source_lang = st.text_input("Source language", value=record.source_lang)
                with col2:
                    target_lang = st.text_input("Target language", value=record.target_lang)
                with col3:
                    tuid = st.text_input("TU ID", value=record.tuid)

                source_text = st.text_area("Source text", value=record.source_text, height=160)
                target_text = st.text_area("Target text", value=record.target_text, height=160)

                submitted = st.form_submit_button("Save Changes")

            if submitted:
                record.source_lang = source_lang.strip()
                record.target_lang = target_lang.strip()
                record.tuid = tuid.strip()
                record.source_text = source_text.strip()
                record.target_text = target_text.strip()
                record.source_length = len(record.source_text)
                record.target_length = len(record.target_text)

                rerun_qa(settings)
                log(f"Edited record {record.record_id}.")
                st.success("Record updated and QA re-run.")

            if st.button("Delete This Record"):
                st.session_state.records = [r for r in st.session_state.records if r.record_id != selected_id]
                rerun_qa(settings)
                log(f"Deleted record {selected_id}.")
                st.success("Record deleted.")
                st.rerun()

    with tab_export:
        records = st.session_state.records

        if not records:
            st.info("No records loaded.")
        else:
            st.subheader("Export Sanitized Assets")

            col1, col2 = st.columns(2)

            with col1:
                tmx_bytes = TMXExporter.to_bytes(records)
                st.download_button(
                    "Download Sanitized TMX",
                    data=tmx_bytes,
                    file_name="sanitized_tmx_export.tmx",
                    mime="application/xml",
                    use_container_width=True,
                )

            with col2:
                xlsx_bytes = XLSXReportExporter.to_bytes(records, st.session_state.stats)
                st.download_button(
                    "Download XLSX QA Report",
                    data=xlsx_bytes,
                    file_name="tmx_sanitizer_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            st.warning("Use exported TMX files carefully. Always test-import into your CAT/TMS environment before replacing production TMs.")

    with tab_logs:
        if not st.session_state.logs:
            st.info("No logs yet.")
        else:
            st.text_area("Log", "\n".join(st.session_state.logs), height=600)


if __name__ == "__main__":
    main()
